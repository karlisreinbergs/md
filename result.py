from openpyxl import Workbook, load_workbook
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0
max_row=ws.max_row
for i in range(2,max_row+1):
    hours=ws['C'+str(i)].value
    rate=ws['B'+str(i)].value
    if (type(hours)!=str and type(rate)!=str):
        salary=float(hours)*float(rate)
        if (salary>3000):
            ws['D'+str(i)].value=salary
            total=total+1
print(total)
wb.close()